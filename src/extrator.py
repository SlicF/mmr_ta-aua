# -*- coding: utf-8 -*-
"""
Extrator de resultados desportivos a partir de ficheiros Excel da Taça UA.

Fluxo principal:
  1. Descarregar (ou localizar) o Excel de resultados.
  2. Selecionar as folhas da época mais recente.
  3. Para cada folha: extrair jogos, datas, faltas e playoffs.
  4. Guardar CSVs por modalidade em docs/output/csv_modalidades/.
"""

import argparse
import hashlib
import json
import logging
import os
import re
import shutil
import sys
import warnings
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
import requests
from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO, format="%(message)s")
warnings.simplefilter(action="ignore", category=FutureWarning)


# ── Constantes ───────────────────────────────────────────────────────────────

_INVALID_TEAM_SUBSTRINGS = [
    "jornada",
    "dia",
    "hora",
    "local",
    "resultado",
    "equipa visitada",
    "equipa visitante",
    "|",
    "pah",
    "pavilhão",
    "pav.",
    "campo",
    "pista",
    "complexo",
    "multiusos",
]

_BANNED_ROW_TOKENS = {
    "vs",
    "v s",
    "v.s.",
    "x",
    "jornada",
    "resultado",
    "equipa visitada",
    "equipa visitante",
    "playoffs",
    "playoff",
    "dia",
    "hora",
    "local",
    "pah",
    "pavilhão",
    "pav.",
}


# ── Mapeador de estágios de playoff ──────────────────────────────────────────


class _StageMapper:
    """Converte texto de estágio em código de jornada, mantendo o contexto.

    O contexto ('E', 'PM', 'LM') é atualizado à medida que novos cabeçalhos
    de secção são encontrados no PDF/Excel.

    Exemplos:
        mapper = _StageMapper("E")
        mapper.map("Quartos de final")  → "E1"
        mapper.map("Meia Final")        → "E2"
        mapper.map("Final")             → "E3"
    """

    def __init__(self, initial_context: Optional[str] = None):
        self.context = initial_context  # 'E', 'PM', 'LM', ou None

    def map(self, stage: str) -> Optional[str]:
        """Mapeia texto de estágio para código de jornada."""
        s = (stage or "").strip().lower()
        if not s:
            return None

        # Atualizar contexto pelo conteúdo do texto
        if any(k in s for k in ["ligu", "ligui"]):
            self.context = "LM"
        elif any(k in s for k in ["manuten", "manutençao", "manutenção", "promo"]):
            self.context = "PM"
        elif "playoff" in s and self.context is None:
            self.context = "E"

        # Mapear pela fase dentro do contexto atual
        if self.context == "PM":
            if s.startswith("meia") or "semi" in s:
                return "PM1"
            if s.startswith("final"):
                return "PM2"
        elif self.context == "LM":
            return "LM"
        elif self.context == "E":
            if s.startswith("quartos"):
                return "E1"
            if s.startswith("meia") or "semi" in s:
                return "E2"
            if ("3" in s and "4" in s) or "3º" in s or "3o" in s:
                return "E3L"
            if s.startswith("final"):
                return "E3"

        # Fallback sem contexto definido
        if s.startswith("quartos"):
            self.context = "E"
            return "E1"
        if s.startswith("meias") or "semi" in s:
            if self.context == "PM":
                return "PM1"
            self.context = "E"
            return "E2"
        if ("3" in s and "4" in s) or "3º" in s or "3o" in s:
            self.context = "E"
            return "E3L"
        if s.startswith("final"):
            if self.context == "PM":
                return "PM2"
            self.context = "E"
            return "E3"

        return None


# ── Funções utilitárias de ficheiros e época ─────────────────────────────────


def get_file_hash(filepath: str) -> Optional[str]:
    """Calcula o hash MD5 de um ficheiro."""
    hash_md5 = hashlib.md5()
    try:
        with open(filepath, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_md5.update(chunk)
        return hash_md5.hexdigest()
    except FileNotFoundError:
        return None


def files_are_identical(file1: str, file2: str) -> bool:
    """Verifica se dois ficheiros são idênticos pelos hashes MD5."""
    return get_file_hash(file1) == get_file_hash(file2)


def extract_season_from_filename(filename: str) -> str:
    """Extrai a época no formato 'YY_YY' do nome do ficheiro."""
    match = re.search(r"(\d{2,4})[_-](\d{2})", filename)
    if match:
        year1, year2 = match.groups()
        if len(year1) == 4:
            year1 = year1[-2:]
        return f"{year1}_{year2}"

    year_match = re.search(r"(\d{4})", filename)
    if year_match:
        return year_match.group(1)[-2:]

    return ""


def normalize_results_url(url: str) -> str:
    """Normaliza links comuns (Google Sheets/Drive, OneDrive) para download direto em XLSX."""
    if not url:
        return url
    try:
        lower = url.lower()

        if "docs.google.com/spreadsheets" in lower:
            m = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", url)
            if m:
                return f"https://docs.google.com/spreadsheets/d/{m.group(1)}/export?format=xlsx"

        if "drive.google.com" in lower:
            m = re.search(r"/file/d/([a-zA-Z0-9-_]+)/", url)
            if m:
                return f"https://drive.google.com/uc?export=download&id={m.group(1)}"
            m = re.search(r"[?&]id=([a-zA-Z0-9-_]+)", url)
            if m:
                return f"https://drive.google.com/uc?export=download&id={m.group(1)}"

        if "1drv.ms" in lower:
            sep = "&" if "?" in url else "?"
            return f"{url}{sep}download=1"

        if "sharepoint.com" in lower or "onedrive.live.com" in lower:
            sep = "&" if "?" in url else "?"
            return f"{url}{sep}download=1"

    except Exception as e:
        logging.error(f"Erro ao normalizar URL '{url}': {e}", exc_info=True)

    return url


def download_results_excel(url: str, dest_dir: Optional[Path] = None) -> Path:
    """Descarrega o ficheiro de resultados e devolve o caminho."""
    dest_dir = dest_dir or Path(".")
    dest_dir.mkdir(parents=True, exist_ok=True)

    basename = Path(re.sub(r"[?#].*$", "", url)).name or "Resultados_Taca_UA.xlsx"
    if not basename.lower().endswith(".xlsx"):
        basename += ".xlsx"

    target = dest_dir / basename
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
    with requests.get(url, headers=headers, stream=True, timeout=60) as r:
        r.raise_for_status()
        with open(target, "wb") as f:
            for chunk in r.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)
    return target


def _parse_season_tokens(text: str) -> Optional[Tuple[int, int]]:
    """Extrai (ano_inicial, ano_final) como inteiros de 4 dígitos."""
    if not text:
        return None
    m = re.search(r"(\d{2,4})[_-](\d{2})", text)
    if not m:
        return None
    y1, y2 = m.groups()
    try:
        y1i, y2i = int(y1[-2:]), int(y2)
        start = 2000 + y1i if y1i < 100 else y1i
        end = 2000 + y2i if y2i < 100 else y2i
        if end < start:
            end += 100
        return (start, end)
    except Exception:
        return None


def detect_latest_season_from_sheet_names(sheet_names: List[str]) -> Optional[str]:
    """Devolve a época mais recente encontrada nos nomes das folhas."""
    best: Optional[Tuple[int, int]] = None
    for name in sheet_names:
        tokens = _parse_season_tokens(str(name))
        if tokens and (best is None or tokens > best):
            best = tokens
    if not best:
        return None
    return f"{str(best[0])[-2:]}_{str(best[1])[-2:]}"


def choose_sheets_for_season(
    sheet_names: List[str], season: Optional[str]
) -> List[str]:
    """Filtra folhas pela época; se nenhuma for encontrada, devolve todas."""
    if not season:
        detected = detect_latest_season_from_sheet_names(sheet_names)
        if not detected:
            return list(sheet_names)
        season = detected

    pattern = re.compile(r"\b" + re.escape(season) + r"\b", re.IGNORECASE)
    selected = [s for s in sheet_names if pattern.search(str(s))]
    return selected if selected else list(sheet_names)


def current_season_token(today: Optional[datetime] = None) -> str:
    """Calcula a época atual no formato 'YY_YY'. Épocas começam em agosto."""
    d = today or datetime.today()
    year, month = d.year, d.month
    if month >= 8:
        y1, y2 = year % 100, (year + 1) % 100
    else:
        y1, y2 = (year - 1) % 100, year % 100
    return f"{y1:02d}_{y2:02d}"


def validate_and_fix_date_for_season(date_val, season: str) -> datetime:
    """Valida e tenta corrigir datas fora do intervalo da época desportiva.

    Intervalo válido: setembro do 1.º ano até junho do 2.º ano.
    Corrige casos comuns de inversão de mês (01↔10, ano errado).
    """
    if not isinstance(date_val, (datetime, pd.Timestamp)):
        return date_val
    if not season or "_" not in season:
        return date_val

    try:
        parts = season.split("_")
        year1, year2 = 2000 + int(parts[0]), 2000 + int(parts[1])
        start_date = datetime(year1, 9, 1)
        end_date = datetime(year2, 6, 30)

        if start_date <= date_val <= end_date:
            return date_val

        # Caso 1: mês 01 de year1 → deveria ser mês 10 de year1
        if date_val.month == 1 and date_val.year == year1:
            corrected = date_val.replace(month=10)
            logging.info(
                f"  [!] Data corrigida: {date_val:%Y-%m-%d} → {corrected:%Y-%m-%d} (jan→out)"
            )
            return corrected

        # Caso 1b: mês 12 de year2 → deveria ser mês 12 de year1
        if date_val.month == 12 and date_val.year == year2:
            corrected = date_val.replace(year=year1)
            if start_date <= corrected <= end_date:
                logging.info(
                    f"  [!] Data corrigida: {date_val:%Y-%m-%d} → {corrected:%Y-%m-%d} (dez year2→year1)"
                )
                return corrected

        # Caso 2: mês 01 de year2 → deveria ser mês 10 de year1
        if date_val.month == 1 and date_val.year == year2:
            corrected = date_val.replace(year=year1, month=10)
            logging.info(
                f"  [!] Data corrigida: {date_val:%Y-%m-%d} → {corrected:%Y-%m-%d} (jan year2→out year1)"
            )
            return corrected

        logging.warning(
            f"  [!] Data fora do intervalo ({start_date:%Y-%m-%d} a {end_date:%Y-%m-%d}): {date_val:%Y-%m-%d}"
        )
        return date_val

    except Exception as e:
        logging.error(f"  [!] Erro ao validar data: {e}")
        return date_val


# ── Processador principal ─────────────────────────────────────────────────────


class ExcelProcessor:
    """Processa ficheiros Excel de resultados desportivos da Taça UA."""

    def __init__(
        self,
        file_path: str,
        output_dir: str = "./docs/output/csv_modalidades",
        season_override: Optional[str] = None,
        sheets_to_process: Optional[List[str]] = None,
    ):
        self.file_path = Path(file_path)
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        self.xls = pd.ExcelFile(file_path)

        # Época
        self.season = season_override or extract_season_from_filename(
            str(self.file_path)
        )
        if not self.season:
            detected = detect_latest_season_from_sheet_names(
                list(map(str, self.xls.sheet_names))
            )
            self.season = detected or current_season_token()

        # Raiz do repositório e sys.path (feito uma vez aqui)
        self._repo_root = Path(__file__).resolve().parents[1]
        _src = self._repo_root / "src"
        if str(_src) not in sys.path:
            sys.path.insert(0, str(_src))

        self._sheets_to_process = sheets_to_process

        self.divisao_pattern = re.compile(r"(\d)\s*(?:ª|º||\.)?\s*DIVIS[ÃAO]+O", re.IGNORECASE)
        self.grupo_pattern = re.compile(r"GRUPO [A-Z]")

        self.base_headers = [
            "Jornada",
            "Dia",
            "Hora",
            "Local",
            "Equipa 1",
            "Golos 1",
            "Golos 2",
            "Equipa 2",
            "Falta de Comparência",
        ]

    # ── Cache de calendários ──────────────────────────────────────────────────

    def _get_calendarios(self):
        """Carrega (e faz cache) calendário PDF e configuração de cursos.

        O carregamento acontece apenas uma vez por instância, reutilizando
        os resultados em _assign_date_placeholders e _assign_venue_placeholders.

        Returns:
            (calendario, config_cursos, calendario_playoffs)
        """
        if hasattr(self, "_cached_cal"):
            return self._cached_cal, self._cached_config, self._cached_cal_playoffs

        from calendario_parser import (
            carregar_calendario_epoca,
            carregar_calendario_playoffs,
            carregar_config_cursos,
            normalizar_nome_equipa,
        )

        self._normalizar = normalizar_nome_equipa
        self._cached_config = carregar_config_cursos(self._repo_root)
        self._cached_cal = carregar_calendario_epoca(
            self.season, repo_root=self._repo_root
        )
        self._cached_cal_playoffs = carregar_calendario_playoffs(
            self.season, repo_root=self._repo_root
        )

        return self._cached_cal, self._cached_config, self._cached_cal_playoffs

    # ── Helpers estáticos de playoff ──────────────────────────────────────────

    @staticmethod
    def _extract_teams(
        row: pd.Series,
        col_home: Optional[str],
        col_away: Optional[str],
        df_columns,
        extra_banned: Optional[set] = None,
    ) -> Tuple[str, str]:
        """Extrai par de equipas de uma linha, com fallback robusto.

        Tenta primeiro pelas colunas identificadas (col_home / col_away).
        Se falhar, varre a linha inteira e escolhe os dois primeiros textos
        relevantes.
        """
        banned = _BANNED_ROW_TOKENS | (extra_banned or set())
        cols_lower = {str(c).lower() for c in df_columns}

        t1 = (
            str(row.get(col_home)).strip()
            if col_home and pd.notna(row.get(col_home))
            else ""
        )
        t2 = (
            str(row.get(col_away)).strip()
            if col_away and pd.notna(row.get(col_away))
            else ""
        )

        t1_valid = ExcelProcessor._is_valid_team(t1)
        t2_valid = ExcelProcessor._is_valid_team(t2)

        if t1 and t2 and t1_valid and t2_valid:
            return t1, t2

        # Se pelo menos um for inválido, tentar fallback para encontrar outros nomes
        texts: List[str] = []
        for c in df_columns:
            val = row.get(c)
            if pd.isna(val):
                continue
            s = str(val).strip()
            if not s:
                continue
            sl, cl = s.lower(), str(c).lower()
            if cl.startswith(("jornada", "dia", "hora", "local")):
                continue
            if "result" in cl or sl in banned or "|" in s:
                continue
            if s.isdigit() and len(s) <= 2:
                continue
            if sl.startswith(("quartos", "meias", "final")) or (
                "3" in sl and "4" in sl
            ):
                continue
            if sl in cols_lower:
                continue

            # Filtrar por validade
            if not ExcelProcessor._is_valid_team(s):
                continue

            if s not in texts:
                texts.append(s)

        # Reconstruir par de equipas
        # Se tínhamos uma válida originalmente, usá-la
        final_t1 = t1 if t1_valid else ""
        final_t2 = t2 if t2_valid else ""

        # Preencher vazios com o que foi encontrado no fallback
        for t in texts:
            if not final_t1 and t != final_t2:
                final_t1 = t
            elif not final_t2 and t != final_t1:
                final_t2 = t

        return final_t1, final_t2

    @staticmethod
    def _extract_lp_number(header_text: str) -> Optional[str]:
        """Extrai o número de jornada de liguilha do cabeçalho."""
        if not header_text:
            return None
        m = re.search(r"jornada\s*(\d+)", header_text, flags=re.I)
        if m:
            return m.group(1)
        if re.match(r"^\d+$", header_text.strip()):
            return header_text.strip()
        return None

    @staticmethod
    def _is_valid_team(name: str) -> bool:
        """Devolve False se o nome da equipa for inválido (cabeçalho, data, etc.)."""
        if not name:
            return False

        # Se for um objeto de data/hora
        if hasattr(name, "year") and hasattr(name, "month"):
            return False

        s = str(name).strip()
        if not s:
            return False

        # Rejeitar tokens de um unico caracter (letras de grupo: A, B, C...)
        if len(s) == 1:
            return False

        # Rejeitar se parecer uma data (YYYY-MM-DD ou DD/MM/YYYY)
        if re.match(r"^\d{4}-\d{2}-\d{2}", s) or re.match(r"^\d{2}/\d{2}/\d{4}", s):
            return False

        # Rejeitar se parecer uma hora (ex: 12h30, 12:30, 12h)
        if re.match(r"^\d{1,2}[h:]\d{0,2}$", s.lower()):
            return False

        nl = s.lower()
        return not any(p in nl for p in _INVALID_TEAM_SUBSTRINGS)

    @staticmethod
    def _find_col(df_columns, *substrings: str) -> Optional[str]:
        """Devolve o primeiro nome de coluna que contenha algum dos substrings."""
        for c in df_columns:
            cl = str(c).lower()
            if any(s in cl for s in substrings):
                return c
        return None

    @staticmethod
    def _to_int_goal(value) -> object:
        """Converte um valor de golo para inteiro ou mantém string com grandes penalidades."""
        if pd.isna(value):
            return None

        # Evitar tratar datas como golos
        if hasattr(value, "year") and hasattr(value, "month"):
            return None

        if isinstance(value, (int, float)):
            if pd.isna(value):
                return None
            try:
                return int(value)
            except Exception:
                return None

        s = str(value).strip()
        if not s:
            return None
        if s.lower() in {"vs", "v.s.", "v s", "x", "nan", "none", "null"}:
            return None

        # Se tiver parênteses (ex. grandes penalidades), retornar como string
        if "(" in s and ")" in s:
            return s

        m = re.match(r"^\d+$", s)
        if not m:
            try:
                return int(float(s))
            except Exception:
                return None
        try:
            return int(s)
        except Exception:
            return None

    @staticmethod
    def _extract_goals(
        row: pd.Series,
        col_home: Optional[str],
        col_away: Optional[str],
        df_columns,
    ) -> Tuple[object, object]:
        """Extrai golos da linha de playoff (ex.: ""2 vs 0"").

        Procura primeiro colunas explícitas de golos/resultado. Se não existir,
        usa as colunas entre equipa visitada e visitante.
        """
        # 1) Tentativa por nomes de coluna explícitos
        col_g1 = ExcelProcessor._find_col(df_columns, "golos 1", "resultado 1")
        col_g2 = ExcelProcessor._find_col(df_columns, "golos 2", "resultado 2")
        if col_g1 and col_g2:
            g1 = ExcelProcessor._to_int_goal(row.get(col_g1))
            g2 = ExcelProcessor._to_int_goal(row.get(col_g2))
            return (g1 if g1 is not None else pd.NA, g2 if g2 is not None else pd.NA)

        # 2) Fallback pelo intervalo entre equipa visitada e visitante
        cols_list = list(df_columns)
        if col_home in cols_list and col_away in cols_list:
            ih = cols_list.index(col_home)
            ia = cols_list.index(col_away)
            if ih < ia:
                between_cols = cols_list[ih + 1 : ia]
                goals: List[int] = []
                for c in between_cols:
                    g = ExcelProcessor._to_int_goal(row.get(c))
                    if g is not None:
                        goals.append(g)
                        if len(goals) == 2:
                            break
                if len(goals) == 2:
                    return goals[0], goals[1]

        return pd.NA, pd.NA

    # ── Células vermelhas (faltas de comparência) ─────────────────────────────

    def is_red_cell(self, cell) -> bool:
        """Verifica se uma célula tem fonte vermelha."""
        if not (cell.font.color and cell.value is not None):
            return False
        if hasattr(cell.font.color, "rgb") and cell.font.color.rgb:
            return str(cell.font.color.rgb).upper() in {
                "FFFF0000",
                "FF0000",
                "FFCC0000",
                "CC0000",
            }
        return False

    def extract_red_cells(self, sheet_name: str) -> Dict[int, str]:
        """Extrai posições de células vermelhas (faltas) de uma folha."""
        wb = load_workbook(self.file_path, data_only=True)
        ws = wb[sheet_name]
        linhas_faltas: Dict[int, str] = {}

        for row in ws.iter_rows():
            for cell in row:
                if (
                    hasattr(cell, "column")
                    and cell.column in (5, 9)
                    and self.is_red_cell(cell)
                ):
                    row_num = cell.row
                    if row_num in linhas_faltas:
                        linhas_faltas[row_num] += f", {cell.value}"
                    else:
                        linhas_faltas[row_num] = str(cell.value)

        return linhas_faltas

    # ── Classificação de jornadas e equipas de playoff ────────────────────────

    def is_playoff_jornada(self, jornada_value) -> bool:
        """Verifica se uma jornada é de playoff (E*, PM*, LM*, MP*, LP*)."""
        if not isinstance(jornada_value, str):
            return False
        j = jornada_value.upper().strip()
        if not j:
            return False
        return (
            j.startswith("E")
            or j.startswith("PM")
            or j.startswith("MP")
            or j.startswith("LM")
            or j.startswith("LP")
        )

    def is_playoff_team_name(self, team_name: str) -> bool:
        """Verifica se o nome é uma legenda de playoff (equipa ainda não definida)."""
        if not isinstance(team_name, str):
            return False
        t = team_name.upper().strip()
        patterns = [
            r"^\d+º CLASS\.",
            r"^VENCEDOR",
            r"^VENCIDO",
            r"^FINALISTA",
            r"^MELHOR",
            r"^PIOR",
        ]
        return any(re.search(p, t) for p in patterns)

    def filter_playoff_games(self, df: pd.DataFrame) -> pd.DataFrame:
        """Remove jogos com legendas de playoff não definidas (exceto jornadas de playoff)."""
        if df.empty:
            return df

        placeholder_mask = df["Equipa 1"].apply(self.is_playoff_team_name) | df[
            "Equipa 2"
        ].apply(self.is_playoff_team_name)
        jornada_is_playoff = df["Jornada"].apply(self.is_playoff_jornada)
        remove_mask = placeholder_mask & ~jornada_is_playoff
        filtered = df[~remove_mask].copy()

        if remove_mask.any():
            logging.info(
                f"  - Removidos {remove_mask.sum()} jogos com legendas não definidas"
            )

        playoff_games = filtered[filtered["Jornada"].apply(self.is_playoff_jornada)]
        if not playoff_games.empty:
            logging.info(
                f"  - Preservados {len(playoff_games)} jogos de playoff: "
                f"{list(playoff_games['Jornada'].unique())}"
            )

        return filtered

    # ── Folhas de PLAYOFFS (Excel) ────────────────────────────────────────────

    def _detect_base_modality_for_playoffs(self, sheet_name: str) -> Optional[str]:
        """Descobre a modalidade base de uma folha '* | PLAYOFFS'."""
        if "|" not in sheet_name:
            return None
        prefix = sheet_name.split("|")[0].strip()

        candidates = [
            name.split("|")[0].strip()
            for name in map(str, self.xls.sheet_names)
            if name != sheet_name
            and name.upper().startswith(prefix.upper() + " ")
            and "|" in name
        ]
        if not candidates:
            return prefix
        candidates.sort(key=len, reverse=True)
        return candidates[0]

    def _parse_playoffs_sheet(self, sheet_name: str) -> Optional[pd.DataFrame]:
        """Transforma folha de PLAYOFFS em DataFrame normalizado."""
        try:
            df_raw = pd.read_excel(self.xls, sheet_name=sheet_name)
        except Exception as e:
            logging.warning(
                f"Não foi possível ler a folha de playoffs '{sheet_name}': {e}"
            )
            return None

        if df_raw.empty:
            return None

        # Contexto inicial a partir do nome da folha
        sheet_lower = sheet_name.lower()
        if any(k in sheet_lower for k in ["ligu", "ligui"]):
            initial_context = "LM"
        elif any(k in sheet_lower for k in ["manuten", "manutenção", "promo"]):
            initial_context = "PM"
        else:
            initial_context = "E"

        cols = df_raw.columns
        col_first = cols[0]
        col_dia = self._find_col(cols, "dia")
        col_hora = self._find_col(cols, "hora")
        col_local = self._find_col(cols, "local")
        col_home = self._find_col(cols, "visitad", "equipa visit", "equipa 1")
        col_away = self._find_col(cols, "visitant", "equipa visitan", "equipa 2")

        if not col_home or not col_away:
            try:
                col_home = col_home or (cols[4] if len(cols) > 4 else None)
                col_away = col_away or (cols[8] if len(cols) > 8 else cols[-1])
            except Exception:
                pass

        # Tokens proibidos específicos desta folha
        sheet_extra_banned = {
            tok.strip() for tok in sheet_name.lower().split() if len(tok) > 3
        }

        mapper = _StageMapper(initial_context)
        rows: List[dict] = []
        current_stage: Optional[str] = None

        for _, row in df_raw.iterrows():
            header_text = (
                str(row.get(col_first, "")).strip()
                if pd.notna(row.get(col_first))
                else ""
            )

            maybe_stage = mapper.map(header_text)
            if maybe_stage:
                current_stage = maybe_stage

            lp_number = (
                self._extract_lp_number(header_text) if current_stage == "LM" else None
            )

            team1, team2 = self._extract_teams(
                row, col_home, col_away, cols, sheet_extra_banned
            )
            golos1, golos2 = self._extract_goals(row, col_home, col_away, cols)

            if not team1 and not team2:
                continue

            # Se uma equipa for inválida (ex: data), substituir por placeholder em vez de saltar a linha
            if not self._is_valid_team(team1) and team1 != "":
                team1 = "Equipa Indefinida"
            if not self._is_valid_team(team2) and team2 != "":
                team2 = "Equipa Indefinida"

            # Descartar se uma das equipas ficou vazia (artefacto de parsing)
            if not team1 or not team2:
                continue

            # Determinar código de jornada
            if current_stage in {"E1", "E2", "E3L", "E3", "PM1", "PM2"}:
                jornada_val = current_stage
            elif current_stage == "LM" and lp_number:
                jornada_val = f"LM{lp_number}"
            else:
                jornada_val = (
                    header_text if self.is_playoff_jornada(header_text) else ""
                )

            rows.append(
                {
                    "Jornada": jornada_val,
                    "Dia": str(row.get(col_dia)) if col_dia else "",
                    "Hora": str(row.get(col_hora)) if col_hora else "",
                    "Local": str(row.get(col_local)) if col_local else "",
                    "Equipa 1": team1,
                    "Golos 1": golos1,
                    "Golos 2": golos2,
                    "Equipa 2": team2,
                    "Falta de Comparência": "",
                }
            )

        if not rows:
            return None

        df_out = pd.DataFrame(rows)
        df_out = df_out[df_out["Jornada"].apply(self.is_playoff_jornada)]
        return df_out.reset_index(drop=True)

    def _extract_playoffs_from_dataframe(
        self, df: pd.DataFrame
    ) -> Optional[pd.DataFrame]:
        """Extrai blocos de playoffs embutidos numa folha regular (Quartos, Meias, Final, …)."""
        if df is None or df.empty:
            return None

        cols = df.columns
        col_first = cols[0]
        col_dia = self._find_col(cols, "dia")
        col_hora = self._find_col(cols, "hora")
        col_local = self._find_col(cols, "local")
        col_home = self._find_col(cols, "visitad", "equipa visitad", "equipa 1")
        col_away = self._find_col(cols, "visitant", "equipa visitan", "equipa 2")

        if not col_home or not col_away:
            try:
                col_home = col_home or (cols[4] if len(cols) > 4 else None)
                col_away = col_away or (cols[-1] if len(cols) > 5 else None)
            except Exception:
                pass

        if not col_home or not col_away:
            return None

        mapper = _StageMapper()
        rows: List[dict] = []
        current_stage: Optional[str] = None

        for _, row in df.iterrows():
            # Detetar mudança de contexto por células com pipe ("|")
            for val in row.values:
                if pd.notna(val) and "|" in str(val):
                    section_text = str(val).lower()
                    if any(k in section_text for k in ["ligu", "ligui"]):
                        mapper.context = "LM"
                    elif any(
                        k in section_text for k in ["manuten", "manutenção", "promo"]
                    ):
                        mapper.context = (
                            "LM"
                            if any(k in section_text for k in ["ligu", "ligui"])
                            else "PM"
                        )
                    elif "playoff" in section_text:
                        mapper.context = "E"
                    break  # só precisa de um match

            header_text = (
                str(row.get(col_first, "")).strip()
                if pd.notna(row.get(col_first))
                else ""
            )

            # Ignorar linhas de cabeçalho de secção (contêm "|")
            if any(pd.notna(v) and "|" in str(v) for v in row.values):
                continue

            maybe_stage = mapper.map(header_text)
            if maybe_stage:
                current_stage = maybe_stage

            lp_number = (
                self._extract_lp_number(header_text) if current_stage == "LM" else None
            )

            team1, team2 = self._extract_teams(row, col_home, col_away, cols)
            golos1, golos2 = self._extract_goals(row, col_home, col_away, cols)

            if not team1 and not team2:
                continue

            # Se uma equipa for inválida (ex: data), substituir por placeholder em vez de saltar a linha
            if not self._is_valid_team(team1) and team1 != "":
                team1 = "Equipa Indefinida"
            if not self._is_valid_team(team2) and team2 != "":
                team2 = "Equipa Indefinida"

            # Descartar se uma das equipas ficou vazia (artefacto de parsing)
            if not team1 or not team2:
                continue

            # Determinar código de jornada
            if current_stage in {"E1", "E2", "E3L", "E3", "PM1", "PM2"}:
                jornada_val = current_stage
            elif current_stage == "LM" and lp_number:
                jornada_val = f"LM{lp_number}"
            else:
                jornada_val = (
                    header_text if self.is_playoff_jornada(header_text) else ""
                )

            if not jornada_val:
                continue

            rows.append(
                {
                    "Jornada": jornada_val,
                    "Dia": str(row.get(col_dia)) if col_dia else "",
                    "Hora": str(row.get(col_hora)) if col_hora else "",
                    "Local": str(row.get(col_local)) if col_local else "",
                    "Equipa 1": team1,
                    "Golos 1": golos1,
                    "Golos 2": golos2,
                    "Equipa 2": team2,
                    "Falta de Comparência": "",
                }
            )

        if not rows:
            return None

        df_out = pd.DataFrame(rows)
        df_out = df_out[df_out["Jornada"].apply(self.is_playoff_jornada)]
        df_out = df_out.drop_duplicates(
            subset=["Jornada", "Equipa 1", "Equipa 2"]
        ).reset_index(drop=True)
        return df_out

    def _append_playoffs_to_target_csv(
        self, base_modality: str, playoffs_df: pd.DataFrame
    ):
        """Anexa linhas de playoffs ao CSV da modalidade alvo."""
        filename = (
            f"{base_modality}_{self.season}.csv"
            if self.season
            else f"{base_modality}.csv"
        )
        target_path = self.output_dir / filename

        if not target_path.exists():
            pd.DataFrame(columns=self.base_headers).to_csv(target_path, index=False)

        try:
            base_df = pd.read_csv(target_path)
        except Exception:
            base_df = pd.DataFrame(columns=self.base_headers)

        base_df = self._coerce_integer_columns(base_df)
        if "Época" in base_df.columns:
            base_df = base_df.drop(columns=["Época"])

        # Remover linhas de playoff ja existentes no CSV base para que
        # a versao recentemente processada (com datas do PDF) as substitua.
        # Evita acumulacao entre execucoes.
        if "Jornada" in base_df.columns:
            base_df = base_df[
                ~base_df["Jornada"].apply(lambda j: self.is_playoff_jornada(str(j)))
            ].copy()

        # Inicializar colunas de metadados se não existirem
        playoffs_df = playoffs_df.copy()
        if "Data_Placeholder" not in playoffs_df.columns:
            playoffs_df["Data_Placeholder"] = False
        if "Fonte_Data" not in playoffs_df.columns:
            playoffs_df["Fonte_Data"] = ""

        # Partilhar o contador de slots entre datas e locais para que
        # cada jogo consuma o slot correcto nas duas operações
        self._shared_playoff_indices: dict = {}
        playoffs_df = self._assign_date_placeholders(playoffs_df, base_modality)
        playoffs_df = self._assign_venue_placeholders(playoffs_df, base_modality)
        del self._shared_playoff_indices

        for c in base_df.columns:
            if c not in playoffs_df.columns:
                playoffs_df[c] = pd.NA if c in ("Golos 1", "Golos 2") else ""
        playoffs_df = playoffs_df[base_df.columns]
        playoffs_df = self._coerce_integer_columns(playoffs_df)

        combined = pd.concat([base_df, playoffs_df], ignore_index=True)
        combined = self._coerce_integer_columns(combined)

        # Deduplicar apenas por identidade do jogo (Jornada + equipas).
        # Não incluir Dia/Hora/Local/Golos para que versões com e sem data
        # não escapem à deduplicação; keep="last" preserva a versão mais
        # recente (com dados do PDF de playoffs, adicionada no final).
        dup_cols = ["Jornada", "Equipa 1", "Equipa 2"]
        combined = combined.drop_duplicates(subset=dup_cols, keep="last")
        combined.to_csv(target_path, index=False)
        logging.info(f"  - Playoffs adicionados ao ficheiro: {target_path}")

    # ── Resultados padrão e desistências ──────────────────────────────────────

    def get_sport_default_score(self, sheet_name: str) -> Tuple[int, int]:
        """Resultado padrão (vencedor, perdedor) por desporto."""
        sheet_upper = sheet_name.upper()
        if "VOLEIBOL" in sheet_upper:
            return (2, 0)
        if "FUTSAL" in sheet_upper or "FUTEBOL" in sheet_upper:
            return (3, 0)
        if "ANDEBOL" in sheet_upper:
            return (15, 0)
        if "BASQUETEBOL" in sheet_upper:
            return (21, 0)
        return (3, 0)

    def _detect_withdrawn_teams(self, df: pd.DataFrame) -> set:
        """Deteta equipas que desistiram (todos os jogos com falta de comparência)."""
        if "Falta de Comparência" not in df.columns:
            return set()

        team_games: Dict[str, Dict[str, int]] = {}

        for _, row in df.iterrows():
            falta = row["Falta de Comparência"]
            has_absence = pd.notna(falta) and falta != ""
            equipas_ausentes = (
                {e.strip() for e in falta.split(",") if e.strip()}
                if has_absence
                else set()
            )

            for col in ("Equipa 1", "Equipa 2"):
                equipa = row[col]
                if equipa not in team_games:
                    team_games[equipa] = {"total": 0, "with_absence": 0}
                team_games[equipa]["total"] += 1
                if equipa in equipas_ausentes:
                    team_games[equipa]["with_absence"] += 1

        withdrawn = {
            team
            for team, stats in team_games.items()
            if stats["total"] > 0 and stats["with_absence"] == stats["total"]
        }
        for team in withdrawn:
            logging.info(
                f"  [!] Equipa desistente: {team} ({team_games[team]['total']} jogos)"
            )
        return withdrawn

    def apply_default_scores(self, df: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
        """Aplica resultados padrão quando uma equipa (não desistente) faltou."""
        if "Falta de Comparência" not in df.columns:
            return df

        golos_vencedor, golos_perdedor = self.get_sport_default_score(sheet_name)
        withdrawn = self._detect_withdrawn_teams(df)

        for idx, row in df.iterrows():
            falta = row["Falta de Comparência"]
            if not (pd.notna(falta) and falta != ""):
                continue

            equipas_faltaram = [e.strip() for e in falta.split(",") if e.strip()]
            if len(equipas_faltaram) != 1:
                continue

            equipa_faltou = equipas_faltaram[0]
            if row["Equipa 1"] in withdrawn or row["Equipa 2"] in withdrawn:
                continue

            mask = df.index == idx
            if equipa_faltou == row["Equipa 1"]:
                df.loc[mask, "Golos 1"] = golos_perdedor
                df.loc[mask, "Golos 2"] = golos_vencedor
            elif equipa_faltou == row["Equipa 2"]:
                df.loc[mask, "Golos 1"] = golos_vencedor
                df.loc[mask, "Golos 2"] = golos_perdedor

        return df

    # ── Divisões e grupos ─────────────────────────────────────────────────────

    def extract_division_number(self, text: str) -> Optional[str]:
        m = self.divisao_pattern.search(text) if isinstance(text, str) else None
        return m.group(1) if m else None

    def extract_group(self, text: str) -> Optional[str]:
        m = self.grupo_pattern.search(str(text)) if text else None
        return m.group(0) if m else None

    def find_divisions_and_groups(
        self, df: pd.DataFrame
    ) -> Tuple[Dict[str, int], Dict[str, int]]:
        """Encontra divisões e grupos no DataFrame e devolve os seus índices de início."""
        primeira_coluna = df.columns[0]
        divisoes: Dict[str, int] = {}
        grupos: Dict[str, int] = {}

        num_div = self.extract_division_number(primeira_coluna)
        if num_div:
            divisoes[num_div] = 0
        grupo = self.extract_group(primeira_coluna)
        if grupo:
            grupos[grupo] = 0

        for idx, valor in enumerate(df[primeira_coluna]):
            if isinstance(valor, str):
                norm_val = valor.upper().replace("Ã", "A").replace("Õ", "O").replace("Ç", "C")
                if "DIVISAO" in norm_val:
                    nd = self.extract_division_number(valor)
                    if nd and nd not in divisoes:
                        divisoes[nd] = idx
                    g = self.extract_group(valor)
                    if g and g not in grupos:
                        grupos[g] = idx
            else:
                g = self.extract_group(str(valor))
                if g and g not in grupos:
                    grupos[g] = idx

        return divisoes, grupos

    def fill_sections(
        self, df: pd.DataFrame, sections: Dict[str, int], column_name: str
    ) -> pd.DataFrame:
        """Preenche a coluna de secção (Divisão ou Grupo) por intervalos."""
        if not sections:
            return df

        df[column_name] = ""
        indices = sorted(sections.items(), key=lambda x: x[1])

        for i, (name, start_idx) in enumerate(indices):
            end_idx = indices[i + 1][1] if i + 1 < len(indices) else len(df)
            clean_name = (
                name.replace("GRUPO ", "").strip() if column_name == "Grupo" else name
            )
            df.loc[start_idx : end_idx - 1, column_name] = clean_name

        return df

    def create_headers(self, has_divisions: bool, has_groups: bool) -> List[str]:
        headers = self.base_headers.copy()
        if has_divisions:
            headers.append("Divisão")
        if has_groups:
            headers.append("Grupo")
        return headers

    # ── Atribuição de datas ───────────────────────────────────────────────────

    def _load_existing_csv_dates(
        self, modality: str
    ) -> Dict[Tuple[str, str, str], Tuple[str, str, bool]]:
        """Carrega datas do CSV anterior para preservar placeholders."""
        filename = (
            f"{modality.upper()}_{self.season}.csv"
            if self.season
            else f"{modality.upper()}.csv"
        )
        csv_path = self.output_dir / filename

        if not csv_path.exists():
            return {}

        try:
            df_old = pd.read_csv(csv_path)
            date_map = {}
            for _, row in df_old.iterrows():
                key = (
                    str(row.get("Equipa 1", "")).strip(),
                    str(row.get("Equipa 2", "")).strip(),
                    str(row.get("Jornada", "")).strip(),
                )
                is_ph = row.get("Data_Placeholder", False)
                if isinstance(is_ph, str):
                    is_ph = is_ph.lower() in ("true", "1", "yes")
                date_map[key] = (row.get("Dia", ""), row.get("Hora", ""), bool(is_ph))
            return date_map
        except Exception as e:
            logging.error(f"  [!] Erro ao carregar datas antigas: {e}")
            return {}

    def _assign_date_placeholders(
        self, df: pd.DataFrame, modality: str
    ) -> pd.DataFrame:
        """Atribui datas com prioridade corrigida:
        Para Playoffs: PDF Playoffs -> Excel -> Placeholder.
        Para Regular: Excel -> PDF Regular -> Placeholder.
        """
        cal, config, cal_playoffs = self._get_calendarios()

        if "Data_Placeholder" not in df.columns:
            df["Data_Placeholder"] = False
        if "Fonte_Data" not in df.columns:
            df["Fonte_Data"] = ""

        import re

        modality_clean = re.sub(
            r"(_25_26|_24_25|\|.*|PLAYOFFS.*)$", "", modality, flags=re.I
        ).strip()

        modality_map = {
            "ANDEBOL MISTO": "ANDEBOL",
            "FUTEBOL DE 7 MASCULINO": "FUTEBOL 7",
        }
        pdf_modality = modality_map.get(modality_clean, modality_clean)

        logging.info(
            f"[PLAYOFFS] Modalidade: {modality} -> pdf_modality: {pdf_modality}"
        )
        if cal_playoffs:
            keys_for_mod = [k for k in cal_playoffs.keys() if k[0] == pdf_modality]
            logging.info(
                f"[PLAYOFFS] Chaves disponíveis para {pdf_modality}: {keys_for_mod}"
            )

        def _lookup_pdf_date(
            equipa1: str, equipa2: str
        ) -> Tuple[Optional[str], Optional[str]]:
            if pdf_modality not in cal:
                return None, None
            e1n = self._normalizar(equipa1, config)
            e2n = self._normalizar(equipa2, config)
            games = cal[pdf_modality]
            for key in ((e1n, e2n), (e2n, e1n)):
                if key in games:
                    v = games[key]
                    if isinstance(v, (tuple, list)) and len(v) >= 2:
                        return v[0], v[1]
            return None, None

        def _parse_date_only(value):
            if pd.isna(value):
                return None
            dt = pd.to_datetime(str(value), errors="coerce")
            return dt.date() if not pd.isna(dt) else None

        def _is_day_month_swap(excel_date, pdf_date) -> bool:
            if not excel_date or not pdf_date or excel_date.year != pdf_date.year:
                return False
            return (
                excel_date.day == pdf_date.month
                and excel_date.month == pdf_date.day
                and excel_date.day <= 12
                and excel_date.month <= 12
            )

        def _swap_day_month(d):
            if not d or d.day > 12 or d.month > 12:
                return None
            try:
                return d.replace(day=d.month, month=d.day)
            except ValueError:
                return None

        games_for_placeholder: List[int] = []
        # Usar contador partilhado (criado em _append_playoffs_to_target_csv)
        # para sincronizar com _assign_venue_placeholders
        playoff_indices = getattr(self, "_shared_playoff_indices", {})

        for idx, row in df.iterrows():
            dia = row.get("Dia", "")
            hora = row.get("Hora", "")
            golos1, golos2 = row.get("Golos 1", ""), row.get("Golos 2", "")
            desistencia = str(row.get("Desistência", "")).strip().upper()
            equipa1 = str(row.get("Equipa 1", "")).strip()
            equipa2 = str(row.get("Equipa 2", "")).strip()
            jornada = str(row.get("Jornada", "")).strip().upper()

            has_result = (
                pd.notna(golos1)
                and str(golos1).strip() != ""
                and pd.notna(golos2)
                and str(golos2).strip() != ""
            )
            has_excel_date = (
                pd.notna(dia)
                and str(dia).strip() != ""
                and pd.notna(hora)
                and str(hora).strip() != ""
            )

            # --- LÓGICA PARA JOGOS DE PLAYOFF (E*, PM*, LM*) ---
            if self.is_playoff_jornada(jornada):
                # Normalizar jornada para busca flexível
                jornada_normalized = jornada.upper().strip()

                # PDF de Playoffs é a fonte de datas para playoffs
                if cal_playoffs:
                    playoff_key = (pdf_modality, jornada_normalized)

                    logging.info(
                        f"[PLAYOFFS] Procurando chave: {playoff_key} em {list(cal_playoffs.keys())[:5]}..."
                    )

                    # Tentar chave exata primeiro
                    if playoff_key in cal_playoffs and cal_playoffs[playoff_key]:
                        idx_in_list = playoff_indices.get(playoff_key, 0)
                        total_slots = len(cal_playoffs[playoff_key])
                        logging.info(
                            f"[PLAYOFFS] Encontrada chave {playoff_key} com {total_slots} slots, idx_atual={idx_in_list}"
                        )

                        if idx_in_list < total_slots:
                            slot = cal_playoffs[playoff_key][idx_in_list]
                            playoff_indices[playoff_key] = idx_in_list + 1
                            df.at[idx, "Dia"] = slot[0]
                            df.at[idx, "Hora"] = slot[1]
                            df.at[idx, "Fonte_Data"] = "Calendário PDF Playoffs"
                            df.at[idx, "Data_Placeholder"] = False
                            logging.info(
                                f"[PLAYOFFS] OK {idx}: {jornada_normalized} -> {slot[0]} {slot[1]}"
                            )
                            continue
                        else:
                            logging.warning(
                                f"[PLAYOFFS] Índice excedido: {idx_in_list} >= {total_slots}, usando último slot"
                            )
                            slot = cal_playoffs[playoff_key][-1]
                            df.at[idx, "Dia"] = slot[0]
                            df.at[idx, "Hora"] = slot[1]
                            df.at[idx, "Fonte_Data"] = "Calendário PDF Playoffs"
                            df.at[idx, "Data_Placeholder"] = False
                            logging.info(
                                f"[PLAYOFFS] ULTIMO SLOT {idx}: {jornada_normalized} -> {slot[0]} {slot[1]}"
                            )
                            continue

                    # Fallback: procurar por nomes de equipas (busca exata)
                    e1_normalized = (
                        self._normalizar(equipa1, config) if self._normalizar else None
                    )
                    e2_normalized = (
                        self._normalizar(equipa2, config) if self._normalizar else None
                    )

                    logging.info(
                        f"[PLAYOFFS] Fallback equipas: {e1_normalized} vs {e2_normalized}"
                    )

                    if e1_normalized and e2_normalized:
                        for key in cal_playoffs:
                            if key[0] != pdf_modality:
                                continue
                            for slot_idx, slot in enumerate(cal_playoffs[key]):
                                local = (
                                    str(slot[2]).lower()
                                    if len(slot) >= 3 and slot[2]
                                    else ""
                                )
                                e1_lower = e1_normalized.lower()
                                e2_lower = e2_normalized.lower()
                                if e1_lower in local or e2_lower in local:
                                    df.at[idx, "Dia"] = slot[0]
                                    df.at[idx, "Hora"] = slot[1]
                                    df.at[idx, "Fonte_Data"] = "Calendário PDF Playoffs"
                                    df.at[idx, "Data_Placeholder"] = False
                                    logging.info(
                                        f"[PLAYOFFS-EQUIPAS] OK {idx}: {jornada_normalized} -> {slot[0]} (local={local})"
                                    )
                                    break
                            else:
                                continue
                            break

                    # Fallback: procurar qualquer jogo disponível nesta fase (sem depender de índice)
                    fase_prefix = (
                        jornada_normalized[:2]
                        if len(jornada_normalized) >= 2
                        else jornada_normalized
                    )
                    logging.info(f"[PLAYOFFS] Fallback fase: {fase_prefix}")

                    for key in cal_playoffs:
                        if key[0] == pdf_modality and key[1].startswith(fase_prefix):
                            available_slots = [
                                i
                                for i, s in enumerate(cal_playoffs[key])
                                if i >= playoff_indices.get(key, 0)
                            ]
                            if available_slots:
                                next_idx = available_slots[0]
                                slot = cal_playoffs[key][next_idx]
                                playoff_indices[key] = next_idx + 1
                            else:
                                slot = cal_playoffs[key][-1]
                            df.at[idx, "Dia"] = slot[0]
                            df.at[idx, "Hora"] = slot[1]
                            df.at[idx, "Fonte_Data"] = "Calendário PDF Playoffs"
                            df.at[idx, "Data_Placeholder"] = False
                            logging.info(
                                f"[PLAYOFFS-FALLBACK] OK {idx}: {jornada_normalized} -> {slot[0]} (chave {key})"
                            )
                            break
                    else:
                        logging.warning(
                            f"[PLAYOFFS] Nenhuma chave encontrada para fase {fase_prefix} em {pdf_modality}"
                        )

                # Se chegou aqui sem encontrar data no PDF, usar placeholder se tiver resultado
                if has_result:
                    games_for_placeholder.append(idx)
                else:
                    df.at[idx, "Data_Placeholder"] = False
                    df.at[idx, "Fonte_Data"] = ""
                continue

            # --- LÓGICA PARA JOGOS REGULARES ---
            # Prioridade 1: Data Excel
            if has_excel_date:
                df.at[idx, "Data_Placeholder"] = False
                if not df.at[idx, "Fonte_Data"]:
                    df.at[idx, "Fonte_Data"] = "Excel"

                if desistencia != "SIM":
                    pdf_data, pdf_hora = _lookup_pdf_date(equipa1, equipa2)
                    if pdf_data:
                        excel_date = _parse_date_only(dia)
                        pdf_date = _parse_date_only(pdf_data)
                        if excel_date and excel_date >= datetime.now().date():
                            df.at[idx, "Dia"] = pdf_data
                            df.at[idx, "Hora"] = pdf_hora or hora
                            df.at[idx, "Fonte_Data"] = "Calendário PDF (Futuro)"
                            continue
                        if _is_day_month_swap(excel_date, pdf_date):
                            df.at[idx, "Dia"] = pdf_data
                            df.at[idx, "Hora"] = pdf_hora or hora
                            df.at[idx, "Fonte_Data"] = "Calendário PDF (corrigido)"
                            continue

                if has_result:
                    excel_date = _parse_date_only(dia)
                    if excel_date and excel_date > datetime.now().date():
                        swapped = _swap_day_month(excel_date)
                        if swapped and swapped <= datetime.now().date():
                            df.at[idx, "Dia"] = swapped.strftime("%Y-%m-%d 00:00:00")
                            df.at[idx, "Fonte_Data"] = "Excel (dia/mês corrigido)"
                continue

            # Prioridade 2: Calendário PDF Regular
            if desistencia == "SIM":
                df.at[idx, "Data_Placeholder"] = False
                df.at[idx, "Fonte_Data"] = ""
                continue

            pdf_data, pdf_hora = _lookup_pdf_date(equipa1, equipa2)
            if pdf_data:
                df.at[idx, "Dia"] = pdf_data
                df.at[idx, "Hora"] = pdf_hora
                df.at[idx, "Fonte_Data"] = "Calendário PDF"
                df.at[idx, "Data_Placeholder"] = False
                continue

            # Prioridade 3: Placeholder se tem resultado
            if has_result:
                games_for_placeholder.append(idx)
            else:
                df.at[idx, "Data_Placeholder"] = False
                df.at[idx, "Fonte_Data"] = ""

        # --- Segunda fase: Atribuir placeholders ---
        if games_for_placeholder:
            jornadas_sem_data = df.loc[games_for_placeholder, "Jornada"].unique()
            jornadas_numericas = sorted(
                ((int(j), j) for j in jornadas_sem_data if str(j).isdigit()),
                key=lambda x: x[0],
            )

            if jornadas_numericas:
                base_date = datetime.now()
                available_hours = list(range(15, 24)) + list(range(0, 3))
                jornada_dates = {}

                for i, (jornada_num, jornada_str) in enumerate(jornadas_numericas):
                    hour = available_hours[i % len(available_hours)]
                    days_ago = jornadas_numericas[-1][0] - jornada_num
                    jornada_date = (base_date - pd.Timedelta(days=days_ago)).replace(
                        hour=hour, minute=0, second=0
                    )
                    jornada_dates[jornada_str] = jornada_date

                for idx in games_for_placeholder:
                    jornada = df.at[idx, "Jornada"]
                    ph_date = jornada_dates.get(jornada, base_date)
                    df.at[idx, "Dia"] = ph_date.strftime("%Y-%m-%d 00:00:00")
                    df.at[idx, "Hora"] = ph_date.strftime("%Hh%M")
                    df.at[idx, "Data_Placeholder"] = True
                    df.at[idx, "Fonte_Data"] = ""

        return df

    def _assign_venue_placeholders(
        self, df: pd.DataFrame, modality: str
    ) -> pd.DataFrame:
        """Atribui locais com prioridade: Excel → Calendário PDF → Playoffs PDF → Placeholder."""
        cal, config, cal_playoffs = self._get_calendarios()

        if "Fonte_Local" not in df.columns:
            df["Fonte_Local"] = ""
        if "Local_Placeholder" not in df.columns:
            df["Local_Placeholder"] = False
        if "Local" not in df.columns:
            df["Local"] = ""

        import re

        modality_clean = re.sub(
            r"(_25_26|_24_25|\|.*|PLAYOFFS.*)$", "", modality, flags=re.I
        ).strip()

        modality_map = {
            "ANDEBOL MISTO": "ANDEBOL",
            "FUTEBOL DE 7 MASCULINO": "FUTEBOL 7",
        }
        pdf_modality = modality_map.get(modality_clean, modality_clean)

        def _normalize_modality(text: str) -> str:
            return " ".join(str(text or "").upper().split())

        def _extract_division_number(value) -> Optional[int]:
            if pd.isna(value):
                return None
            m = re.search(r"(\d+)", str(value).strip())
            return int(m.group(1)) if m else None

        def _venue_placeholder(modality_name: str, division_value) -> str:
            norm = _normalize_modality(modality_name)
            div_num = _extract_division_number(division_value)
            if "FUTEBOL DE 7" in norm:
                return "Sintético"
            if "FUTSAL FEMININO" in norm:
                return "Caixa UA"
            if "FUTSAL MASCULINO" in norm:
                return "PAH (Aristides Hall)" if div_num == 2 else "Caixa UA"
            if any(s in norm for s in ("BASQUETEBOL", "VOLEIBOL", "ANDEBOL")):
                return "PAH (Aristides Hall)"
            return ""

        def _clean_text(value) -> str:
            if pd.isna(value):
                return ""
            text = str(value).strip()
            return "" if text.lower() in ("nan", "none", "null") else text

        def _lookup_pdf_venue(equipa1: str, equipa2: str) -> Optional[str]:
            if pdf_modality not in cal:
                return None
            e1n = self._normalizar(equipa1, config)
            e2n = self._normalizar(equipa2, config)
            games = cal[pdf_modality]
            for key in ((e1n, e2n), (e2n, e1n)):
                if key in games:
                    v = games[key]
                    if isinstance(v, (tuple, list)) and len(v) >= 3:
                        local = str(v[2]).strip()
                        if local:
                            return local
            return None

        # Usar contador partilhado (criado em _append_playoffs_to_target_csv)
        # para continuar do ponto onde _assign_date_placeholders parou
        playoff_indices = getattr(self, "_shared_playoff_indices", {})

        for idx, row in df.iterrows():
            local_excel = _clean_text(row.get("Local", ""))
            golos1, golos2 = row.get("Golos 1", ""), row.get("Golos 2", "")
            has_result = (
                pd.notna(golos1)
                and str(golos1).strip() != ""
                and pd.notna(golos2)
                and str(golos2).strip() != ""
            )
            has_date = _clean_text(row.get("Dia", "")) != ""
            equipa1 = str(row.get("Equipa 1", "")).strip()
            equipa2 = str(row.get("Equipa 2", "")).strip()
            jornada = str(row.get("Jornada", "")).strip()
            divisao = row.get("Divisão", row.get("Divisao", ""))

            # Prioridade 1: Local do Excel
            if local_excel:
                df.at[idx, "Local"] = local_excel
                df.at[idx, "Fonte_Local"] = "Excel"
                df.at[idx, "Local_Placeholder"] = False
                continue

            # Prioridade 2: Local do calendário PDF playoffs
            if self.is_playoff_jornada(jornada) and cal_playoffs:
                playoff_key = (pdf_modality, jornada)
                if playoff_key in cal_playoffs and cal_playoffs[playoff_key]:
                    idx_in_list = playoff_indices.get(playoff_key, 0)
                    if idx_in_list < len(cal_playoffs[playoff_key]):
                        slot_local = cal_playoffs[playoff_key][idx_in_list][2]
                        playoff_indices[playoff_key] = idx_in_list + 1
                        if slot_local:
                            df.at[idx, "Local"] = slot_local
                            df.at[idx, "Fonte_Local"] = "Calendário PDF Playoffs"
                            df.at[idx, "Local_Placeholder"] = False
                            continue

            # Prioridade 3: Local do calendário PDF regular
            pdf_local = _lookup_pdf_venue(equipa1, equipa2)
            if pdf_local:
                df.at[idx, "Local"] = pdf_local
                df.at[idx, "Fonte_Local"] = "Calendário PDF"
                df.at[idx, "Local_Placeholder"] = False
                continue

            # Prioridade 4: Placeholder quando há resultado ou data agendada
            if has_result or has_date:
                placeholder_local = _venue_placeholder(modality, divisao)
                if placeholder_local:
                    df.at[idx, "Local"] = placeholder_local
                    df.at[idx, "Fonte_Local"] = "Placeholder"
                    df.at[idx, "Local_Placeholder"] = True
                else:
                    df.at[idx, "Fonte_Local"] = ""
                    df.at[idx, "Local_Placeholder"] = False
            else:
                df.at[idx, "Fonte_Local"] = ""
                df.at[idx, "Local_Placeholder"] = False

        return df

    # ── Limpeza e ordenação do DataFrame ─────────────────────────────────────

    def clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """Remove linhas inválidas, preenche jornadas e converte golos para Int64."""
        df = df.dropna(how="all").reset_index(drop=True)

        primeira_coluna = df.columns[0]
        jornada_numerica = pd.to_numeric(df[primeira_coluna], errors="coerce")
        df = df[df[primeira_coluna].isna() | jornada_numerica.notna()].copy()
        df[primeira_coluna] = jornada_numerica.loc[df.index].astype("Int64")

        df = df[~((df["Equipa 1"] == 0) | (df["Equipa 2"] == 0))]
        df["Jornada"] = df["Jornada"].ffill()

        for col in ("Golos 1", "Golos 2"):
            if col in df.columns:
                def clean_goal(x):
                    if pd.isna(x) or x == "":
                        return pd.NA
                    s = str(x).strip()
                    if "(" in s and ")" in s:
                        return s
                    try:
                        return int(float(s))
                    except (ValueError, TypeError):
                        return pd.NA
                df[col] = df[col].apply(clean_goal)

        return df

    def adjust_journeys(self, df: pd.DataFrame) -> pd.DataFrame:
        """Ajusta números de jornada para evitar duplicações.

        Linhas de playoff (jornada E*, PM*, LM*\u2026) não são alteradas \u2014
        a sua jornada já é um código de fase, não um número sequencial.
        """
        aparicoes: set = set()

        def ajustar_jornada(row):
            j = row["Jornada"]
            # Não modificar jornadas de playoff
            if self.is_playoff_jornada(str(j)):
                return j
            e1, e2 = row["Equipa 1"], row["Equipa 2"]
            if (j, e1) in aparicoes or (j, e2) in aparicoes:
                return j + 1
            aparicoes.add((j, e1))
            aparicoes.add((j, e2))
            return j

        df["Jornada"] = df.apply(ajustar_jornada, axis=1)
        df["Jornada"] = (
            df["Jornada"].astype(str).str.replace(".1", " (2ª)", regex=False)
        )
        return df

    def sort_by_datetime(self, df: pd.DataFrame, modality: str = "") -> pd.DataFrame:
        """Ordena por data/hora. Se modality fornecida, atribui datas e locais antes."""
        if modality:
            df = self._assign_date_placeholders(df, modality)
            df = self._assign_venue_placeholders(df, modality)

        def parse_data_hora(row):
            dia, hora = row["Dia"], row["Hora"]
            if pd.isna(dia):
                return pd.Timestamp.max
            try:
                dt = pd.to_datetime(
                    f"{dia} {hora if not pd.isna(hora) else '00:00'}", errors="coerce"
                )
                if pd.isna(dt):
                    return pd.Timestamp.max
                # Horas da madrugada (0h-1h) ordenam depois da meia-noite
                if dt.time() < pd.to_datetime("01:00").time():
                    dt += pd.Timedelta(hours=24)
                return dt
            except Exception:
                return pd.Timestamp.max

        def parse_jornada_sort(val):
            if isinstance(val, (int, float)):
                return int(val)
            if isinstance(val, str):
                m = re.match(r"^(\d+)", val.strip())
                if m:
                    return int(m.group(1))
            return 10**9

        def parse_divisao_sort(val):
            if pd.isna(val):
                return 10**6
            m = re.search(r"(\d+)", str(val))
            return int(m.group(1)) if m else 10**6

        def parse_grupo_sort(val):
            if pd.isna(val):
                return 10**6
            v = str(val).strip().upper()
            if v and v[0].isalpha():
                return ord(v[0]) - ord("A") + 1
            m = re.search(r"(\d+)", v)
            return int(m.group(1)) if m else 10**6

        df["DataHoraSort"] = df.apply(parse_data_hora, axis=1)
        df["JornadaSort"] = df["Jornada"].apply(parse_jornada_sort)
        df["DivisaoSort"] = (
            df["Divisão"].apply(parse_divisao_sort)
            if "Divisão" in df.columns
            else 10**6
        )
        df["GrupoSort"] = (
            df["Grupo"].apply(parse_grupo_sort) if "Grupo" in df.columns else 10**6
        )

        df = df.sort_values(
            [
                "DataHoraSort",
                "JornadaSort",
                "DivisaoSort",
                "GrupoSort",
                "Equipa 1",
                "Equipa 2",
            ],
            ascending=True,
        )
        df = df.drop(
            columns=["DataHoraSort", "JornadaSort", "DivisaoSort", "GrupoSort"]
        )
        return df

    def finalize_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """Últimas limpezas e reordenação de colunas."""
        colunas_principais = [
            "Dia",
            "Hora",
            "Local",
            "Equipa 1",
            "Golos 1",
            "Golos 2",
            "Equipa 2",
        ]
        df = df.dropna(subset=colunas_principais, how="all")

        for col, default in [
            ("Data_Placeholder", False),
            ("Local_Placeholder", False),
            ("Fonte_Local", ""),
            ("Falta de Comparência", ""),
        ]:
            if col not in df.columns:
                df[col] = default

        # Reordenar: colunas principais → metadados → Falta de Comparência
        meta = ["Fonte_Data", "Fonte_Local", "Data_Placeholder", "Local_Placeholder"]
        colunas = [c for c in df.columns if c not in meta + ["Falta de Comparência"]]
        colunas += [c for c in meta if c in df.columns]
        colunas.append("Falta de Comparência")
        df = df[colunas]

        return self._coerce_integer_columns(df)

    def _coerce_integer_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """Garante que golos e divisões ficam como Int64 no CSV, exceto se contiverem grandes penalidades."""
        df = df.copy()
        def coerce_goal(val):
            if pd.isna(val) or val == "":
                return pd.NA
            s = str(val).strip()
            if "(" in s and ")" in s:
                return s
            try:
                return int(float(s))
            except (ValueError, TypeError):
                return pd.NA

        for col in ("Golos 1", "Golos 2"):
            if col in df.columns:
                df[col] = df[col].apply(coerce_goal)

        for col in ("Divisão", "Divisao"):
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce").astype("Int64")
        return df

    # ── Processamento por folha ───────────────────────────────────────────────

    def process_sheet(self, sheet_name: str) -> bool:
        """Processa uma folha do Excel e guarda o CSV correspondente."""
        if sheet_name == "CASTIGOS":
            return False

        logging.info(f"A processar a folha: {sheet_name}")

        linhas_faltas = self.extract_red_cells(sheet_name)

        df = pd.read_excel(
            self.xls, sheet_name=sheet_name, usecols=[0, 1, 2, 3, 4, 5, 7, 8]
        )
        df["Falta de Comparência"] = ""

        # Procurar se há divisões na folha (scaneando a primeira coluna e cabeçalho)
        def _has_divisao_text(text):
            if not isinstance(text, str):
                return False
            norm = text.upper().replace("Ã", "A").replace("Õ", "O").replace("Ç", "C")
            return "DIVISAO" in norm

        has_div = _has_divisao_text(df.columns[0]) or any(
            _has_divisao_text(val) for val in df.iloc[:, 0]
        )

        # Validar datas fora do intervalo da época
        dia_col = df.columns[1]
        if self.season:
            df[dia_col] = df[dia_col].apply(
                lambda x: (
                    validate_and_fix_date_for_season(x, self.season)
                    if pd.notna(x)
                    else x
                )
            )

        for row_idx, value in linhas_faltas.items():
            df_idx = row_idx - 2
            if 0 <= df_idx < len(df):
                df.at[df_idx, "Falta de Comparência"] = value

        df = df.dropna(how="all").reset_index(drop=True)

        if has_div:
            divisoes, grupos = self.find_divisions_and_groups(df)
            if divisoes and not any(start_idx == 0 for start_idx in divisoes.values()):
                if "1" not in divisoes:
                    divisoes["1"] = 0
            df = self.fill_sections(df, divisoes, "Divisão")
            df = self.fill_sections(df, grupos, "Grupo")
            headers = self.create_headers(bool(divisoes), bool(grupos))
        else:
            _, grupos = self.find_divisions_and_groups(df)
            df = self.fill_sections(df, grupos, "Grupo")
            headers = self.create_headers(False, bool(grupos))

        df = df.iloc[:, : len(headers)]
        df.columns = headers

        # Extrair playoffs embutidos DEPOIS do rename de colunas
        playoffs_df_embedded = self._extract_playoffs_from_dataframe(df.copy())

        # Remover do df principal todas as linhas que fazem parte de blocos
        # de playoff embutidos, ANTES do clean_dataframe. Se o fizermos
        # depois, o ffill da jornada ja lhes atribuiu um numero regular
        # (ex: 4) e is_playoff_jornada nao as remove.
        # Estrategia: percorrer a primeira coluna; assim que encontramos
        # um cabecalho de fase (mapeavel pelo _StageMapper), todas as linhas
        # seguintes sao de playoff e devem ser excluidas do df regular.
        col_jornada = df.columns[0]
        _playoff_stage_seen = False
        _playoff_row_mask = []
        _mapper_temp = _StageMapper()
        for _, row in df.iterrows():
            cell = str(row.get(col_jornada, "")).strip()
            if _mapper_temp.map(cell):
                _playoff_stage_seen = True
            _playoff_row_mask.append(_playoff_stage_seen)
        df = df[~pd.Series(_playoff_row_mask, index=df.index)].copy()

        df = self.clean_dataframe(df)
        df = self.adjust_journeys(df)
        df = self.sort_by_datetime(df, modality=sheet_name)
        df = self.apply_default_scores(df, sheet_name)
        df = self.filter_playoff_games(df)

        df = self.finalize_dataframe(df)

        if "Época" in df.columns:
            df = df.drop(columns=["Época"])
        df = self._coerce_integer_columns(df)

        filename = (
            f"{sheet_name}_{self.season}.csv" if self.season else f"{sheet_name}.csv"
        )
        output_file = self.output_dir / filename
        df.to_csv(output_file, index=False)
        logging.info(f"Folha '{sheet_name}' processada → '{output_file}'")

        # Só anexar playoffs embutidos se não existir folha dedicada de PLAYOFFS
        # para esta modalidade. Se existir, o segundo loop de process_all_sheets
        # trata disso, e não queremos duplicar.
        dedicated = getattr(self, "_modalities_with_dedicated_playoffs", set())
        # Comparar pelo prefixo da folha (antes do "|") porque dedicated
        # guarda o prefixo devolvido por _detect_base_modality_for_playoffs
        sheet_prefix = (
            sheet_name.split("|")[0].strip() if "|" in sheet_name else sheet_name
        )
        if (
            playoffs_df_embedded is not None
            and not playoffs_df_embedded.empty
            and sheet_prefix not in dedicated
        ):
            expected_cols = self.base_headers
            for c in expected_cols:
                if c not in playoffs_df_embedded.columns:
                    playoffs_df_embedded[c] = (
                        pd.NA if c in ("Golos 1", "Golos 2") else ""
                    )
            playoffs_df_embedded = playoffs_df_embedded[expected_cols]
            self._append_playoffs_to_target_csv(sheet_name, playoffs_df_embedded)

        return True

    def process_all_sheets(self):
        """Processa todas as folhas: regulares primeiro, PLAYOFFS depois."""
        processed_count = 0
        target_sheets = self._sheets_to_process or list(map(str, self.xls.sheet_names))

        # Pre-calcular quais modalidades base têm folha dedicada de PLAYOFFS.
        # Para essas, o process_sheet não deve fazer o append dos playoffs
        # embutidos (evitar dupla inserção).
        self._modalities_with_dedicated_playoffs: set = set()
        for sheet in target_sheets:
            if "PLAYOFFS" in sheet.upper():
                bm = self._detect_base_modality_for_playoffs(sheet)
                if bm:
                    self._modalities_with_dedicated_playoffs.add(bm)

        # 1) Folhas regulares
        for sheet in target_sheets:
            if "PLAYOFFS" in sheet.upper():
                continue
            if self.process_sheet(sheet):
                processed_count += 1

        # 2) Folhas de PLAYOFFS → anexar ao CSV da modalidade base
        for sheet in target_sheets:
            if "PLAYOFFS" not in sheet.upper():
                continue

            base_modality = self._detect_base_modality_for_playoffs(sheet)
            if not base_modality:
                logging.warning(
                    f"Não foi possível determinar modalidade base para '{sheet}'. Ignorado."
                )
                continue

            df_playoffs = self._parse_playoffs_sheet(sheet)
            if df_playoffs is None or df_playoffs.empty:
                logging.warning(f"Folha de playoffs '{sheet}' sem linhas válidas.")
                continue

            expected_cols = self.base_headers
            for c in expected_cols:
                if c not in df_playoffs.columns:
                    df_playoffs[c] = pd.NA if c in ("Golos 1", "Golos 2") else ""
            df_playoffs = df_playoffs[expected_cols]
            self._append_playoffs_to_target_csv(base_modality, df_playoffs)

        logging.info(
            f"\nProcessamento concluído! {processed_count} folhas processadas"
            + (
                " e playoffs anexados."
                if any("PLAYOFFS" in s.upper() for s in target_sheets)
                else "."
            )
        )


# ── Ponto de entrada ─────────────────────────────────────────────────────────


def main():
    repo_root = Path(__file__).resolve().parents[1]

    # argumentos de linha de comando
    parser = argparse.ArgumentParser(description="Extrator de resultados Taça UA")
    parser.add_argument(
        "-f",
        "--force",
        action="store_true",
        help="Forçar reprocessamento mesmo se o ficheiro não mudou",
    )
    args = parser.parse_args()

    # 1) Ler URL de resultados de config.json ou variável de ambiente
    config_url: Optional[str] = None
    config_path = repo_root / "docs" / "config" / "config.json"
    if config_path.exists():
        try:
            with open(config_path, "r", encoding="utf-8") as f:
                config_url = json.load(f).get("results_url")
        except Exception as e:
            logging.warning(f"Não foi possível ler config.json: {e}")

    config_url = config_url or os.environ.get("RESULTS_URL")

    downloaded_file: Optional[Path] = None
    season_detected: Optional[str] = None

    # 2) Descarregar ficheiro se URL disponível
    if config_url:
        try:
            url = normalize_results_url(config_url)
            data_dir = repo_root / "data"
            downloaded_file = download_results_excel(url, dest_dir=data_dir)
            logging.info(f"Documento descarregado para: {downloaded_file}")

            xls_temp = pd.ExcelFile(str(downloaded_file))
            season_detected = detect_latest_season_from_sheet_names(
                list(map(str, xls_temp.sheet_names))
            )
            try:
                xls_temp.close()
            except Exception:
                pass

            if not season_detected:
                season_detected = (
                    extract_season_from_filename(downloaded_file.name)
                    or current_season_token()
                )

            target_name = f"Resultados Taça UA {season_detected}.xlsx"
            target_path = (repo_root / "data") / target_name
            if downloaded_file.name != target_name:
                try:
                    if target_path.exists():
                        target_path.unlink()
                    downloaded_file.rename(target_path)
                    downloaded_file = target_path
                except Exception as e:
                    logging.warning(f"Não foi possível renomear: {e}")
                    try:
                        shutil.copy2(downloaded_file, target_path)
                        downloaded_file = target_path
                    except Exception as e2:
                        logging.warning(f"Fallback de cópia falhou: {e2}")

        except Exception as e:
            logging.error(f"Erro ao descarregar o documento: {e}")
            downloaded_file = None

    # 3) Determinar caminho do ficheiro a processar
    if downloaded_file and downloaded_file.exists():
        file_path = str(downloaded_file)
    else:
        default_local = (
            repo_root / "data" / f"Resultados Taça UA {current_season_token()}.xlsx"
        )
        if default_local.exists():
            file_path = str(default_local)
        else:
            candidates = sorted(
                list((repo_root / "data").glob("Resultados Taça UA*.xlsx"))
                + list(repo_root.glob("Resultados Taça UA*.xlsx")),
                key=lambda p: p.stat().st_mtime,
                reverse=True,
            )
            if not candidates:
                logging.error("Nenhum ficheiro local encontrado e URL não disponível.")
                logging.error("Defina 'results_url' em config.json ou RESULTS_URL.")
                return
            file_path = str(candidates[0])

    # 4) Backup e verificação de mudanças
    season_for_backup = (
        season_detected
        or extract_season_from_filename(Path(file_path).name)
        or current_season_token()
    )
    backup_file = str(
        repo_root / "data" / f"backup_Resultados Taça UA {season_for_backup}.xlsx"
    )

    if os.path.exists(backup_file) and files_are_identical(file_path, backup_file):
        if not args.force:
            logging.info(
                "Ficheiro Excel não mudou desde a última execução. Nada a processar."
            )
            if os.getenv("GITHUB_OUTPUT"):
                with open(os.environ["GITHUB_OUTPUT"], "a") as fh:
                    print("data_changed=false", file=fh)
            return
        else:
            logging.info(
                "Ficheiro Excel igual ao backup, mas forçando reprocessamento (--force)."
            )

    logging.info("Ficheiro alterado ou primeira execução. A processar...")

    try:
        shutil.copy2(file_path, backup_file)
        logging.info(f"Backup criado: {backup_file}")
    except Exception as e:
        logging.warning(f"Não foi possível criar backup: {e}")

    # 5) Selecionar folhas e processar
    xls_all = pd.ExcelFile(file_path)
    if not season_detected:
        season_detected = detect_latest_season_from_sheet_names(
            list(map(str, xls_all.sheet_names))
        )

    sheets_to_process = choose_sheets_for_season(
        list(map(str, xls_all.sheet_names)), season_detected
    )

    processor = ExcelProcessor(
        file_path,
        output_dir=str(repo_root / "docs" / "output" / "csv_modalidades"),
        season_override=season_detected,
        sheets_to_process=sheets_to_process,
    )
    processor.process_all_sheets()

    if os.getenv("GITHUB_OUTPUT"):
        with open(os.environ["GITHUB_OUTPUT"], "a") as fh:
            print("data_changed=true", file=fh)


if __name__ == "__main__":
    main()
