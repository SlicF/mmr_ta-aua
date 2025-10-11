import pandas as pd
import warnings
import os
import re
import shutil
import hashlib
import json
import logging
from datetime import datetime
import requests
from openpyxl import load_workbook
from pathlib import Path
from typing import Dict, Tuple, Optional, List

warnings.simplefilter(action="ignore", category=FutureWarning)


def get_file_hash(filepath: str) -> Optional[str]:
    """Calcula o hash MD5 de um arquivo"""
    hash_md5 = hashlib.md5()
    try:
        with open(filepath, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_md5.update(chunk)
        return hash_md5.hexdigest()
    except FileNotFoundError:
        return None


def files_are_identical(file1: str, file2: str) -> bool:
    """Verifica se dois arquivos são idênticos comparando seus hashes"""
    return get_file_hash(file1) == get_file_hash(file2)


def extract_season_from_filename(filename: str) -> str:
    """Extrai a época/temporada do nome do arquivo Excel"""
    # Procura padrões como "24_25", "2024_25", "24-25", "2024-25", etc.
    import re

    # Padrão para capturar anos no formato XX_XX ou XXXX_XX
    pattern = r"(\d{2,4})[_-](\d{2})"
    match = re.search(pattern, filename)

    if match:
        year1, year2 = match.groups()
        # Se o primeiro ano tem 4 dígitos, usa os últimos 2
        if len(year1) == 4:
            year1 = year1[-2:]
        return f"{year1}_{year2}"

    # Se não encontrar o padrão, tenta extrair apenas o ano
    year_pattern = r"(\d{4})"
    year_match = re.search(year_pattern, filename)
    if year_match:
        year = year_match.group(1)
        return year[-2:]  # Retorna os últimos 2 dígitos

    return ""  # Retorna string vazia se não encontrar nada


def normalize_results_url(url: str) -> str:
    """Normaliza links comuns (Google Sheets/Drive, OneDrive/SharePoint) para download direto em XLSX quando possível."""
    if not url:
        return url

    try:
        lower = url.lower()

        # Google Sheets -> exportar como xlsx
        if "docs.google.com/spreadsheets" in lower:
            # Formato: https://docs.google.com/spreadsheets/d/<ID>/edit#... -> export?format=xlsx
            m = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", url)
            if m:
                file_id = m.group(1)
                return f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=xlsx"

        # Google Drive file -> uc?export=download
        if "drive.google.com" in lower:
            m = re.search(r"/file/d/([a-zA-Z0-9-_]+)/", url)
            if m:
                file_id = m.group(1)
                return f"https://drive.google.com/uc?export=download&id={file_id}"
            m = re.search(r"[?&]id=([a-zA-Z0-9-_]+)", url)
            if m:
                file_id = m.group(1)
                return f"https://drive.google.com/uc?export=download&id={file_id}"

        # OneDrive curto 1drv.ms -> ?download=1
        if "1drv.ms" in lower:
            return url + ("&download=1" if "?" in url else "?download=1")

        # OneDrive/SharePoint -> adicionar download=1
        if "sharepoint.com" in lower or "onedrive.live.com" in lower:
            return url + ("&download=1" if "?" in url else "?download=1")

    except Exception as e:
        logging.error(f"Erro ao normalizar URL '{url}': {e}", exc_info=True)

    return url


def download_results_excel(url: str, dest_dir: Optional[Path] = None) -> Path:
    """Descarrega o ficheiro de resultados para o diretório atual (ou dado) e devolve o caminho."""
    dest_dir = dest_dir or Path(".")
    dest_dir.mkdir(parents=True, exist_ok=True)

    # Nome base a partir do URL
    basename = Path(re.sub(r"[?#].*$", "", url)).name or "Resultados_Taca_UA.xlsx"
    # Garantir extensão .xlsx
    if not basename.lower().endswith(".xlsx"):
        basename += ".xlsx"

    target = dest_dir / basename

    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
    with requests.get(url, headers=headers, stream=True, timeout=60) as r:
        r.raise_for_status()
        with open(target, "wb") as f:
            for chunk in r.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)

    return target


def _parse_season_tokens(text: str) -> Optional[Tuple[int, int]]:
    """Extrai tokens de época do texto, devolvendo (ano_inicial, ano_final) em forma completa (ex.: 2024, 2025)."""
    if not text:
        return None
    m = re.search(r"(\d{2,4})[_-](\d{2})", text)
    if not m:
        return None
    y1, y2 = m.groups()
    try:
        y1i = int(y1[-2:])  # usar últimos 2 dígitos se vier 4
        y2i = int(y2)
        # Normalizar para base 2000 para ordenação consistente
        start = 2000 + y1i if y1i < 100 else y1i
        end = 2000 + y2i if y2i < 100 else y2i
        # Se o intervalo de dois dígitos cruza o século (ex.: 99_00), somar 100 ao ano final
        # Regra clara: quando end < start, end += 100
        if end < start:
            end += 100
        return (start, end)
    except Exception:
        return None


def detect_latest_season_from_sheet_names(sheet_names: List[str]) -> Optional[str]:
    """Deteta a época mais recente com base nos nomes das folhas (ex.: '... 24_25')."""
    best: Optional[Tuple[int, int]] = None
    for name in sheet_names:
        tokens = _parse_season_tokens(str(name))
        if tokens:
            if best is None or tokens > best:
                best = tokens
    if not best:
        return None
    # Voltar a formato curto 24_25
    y1_short = str(best[0])[-2:]
    y2_short = str(best[1])[-2:]
    return f"{y1_short}_{y2_short}"


def choose_sheets_for_season(
    sheet_names: List[str], season: Optional[str]
) -> List[str]:
    """Filtra folhas a processar pela época. Se nenhuma época for encontrada, devolve todas."""
    if not season:
        # Ver se pelo menos alguma folha tem época; se sim, escolher a mais recente
        detected = detect_latest_season_from_sheet_names(sheet_names)
        if not detected:
            return list(sheet_names)
        season = detected

    # Usar regex com word boundaries para evitar falsos positivos
    pattern = re.compile(r"\b" + re.escape(season) + r"\b", re.IGNORECASE)
    selected = [s for s in sheet_names if pattern.search(str(s))]
    # Se não encontrar nenhuma com a época explícita, processar todas (compatibilidade)
    return selected if selected else list(sheet_names)


def current_season_token(today: Optional[datetime] = None) -> str:
    """Calcula a época atual no formato 'YY_YY' com base na data atual.

    Regra: épocas começam em agosto. De ago a dez -> ano_atual_ano+1; de jan a jul -> ano-1_ano.
    """
    d = today or datetime.today()
    year = d.year
    month = d.month
    if month >= 8:
        y1 = year % 100
        y2 = (year + 1) % 100
    else:
        y1 = (year - 1) % 100
        y2 = year % 100
    return f"{y1:02d}_{y2:02d}"


class ExcelProcessor:
    """Classe para processar ficheiros Excel de resultados desportivos."""

    def __init__(
        self,
        file_path: str,
        output_dir: str = "csv_modalidades",
        season_override: Optional[str] = None,
        sheets_to_process: Optional[List[str]] = None,
    ):
        self.file_path = Path(file_path)
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        self.xls = pd.ExcelFile(file_path)

        # Extrai a época do nome do arquivo
        self.season = season_override or extract_season_from_filename(
            str(self.file_path)
        )

        # Lista de folhas alvo (opcional)
        self._sheets_to_process = sheets_to_process

        # Padrões regex compilados para melhor performance
        self.divisao_pattern = re.compile(r"(\d)ª DIVISÃO")
        self.grupo_pattern = re.compile(r"GRUPO [A-Z]")

        # Cabeçalhos padrão
        self.base_headers = [
            "Jornada",
            "Dia",
            "Hora",
            "Local",
            "Equipa 1",
            "Golos 1",
            "Golos 2",
            "Equipa 2",
            "Falta de Comparência",
        ]

    def is_red_cell(self, cell) -> bool:
        """Verifica se uma célula tem fonte vermelha."""
        if not (cell.font.color and cell.value is not None):
            return False

        if hasattr(cell.font.color, "rgb") and cell.font.color.rgb:
            color = str(cell.font.color.rgb).upper()
            return color in ["FFFF0000", "FF0000", "FFCC0000", "CC0000"]
        return False

    def extract_red_cells(self, sheet_name: str) -> Dict[int, str]:
        """Extrai células vermelhas de uma folha."""
        wb = load_workbook(self.file_path, data_only=True)
        ws = wb[sheet_name]
        linhas_faltas = {}

        for row in ws.iter_rows():
            for cell in row:
                # Procura apenas nas colunas E e I (colunas 5 e 9, das equipas)
                # Verifica se não é uma célula mesclada e se está nas colunas corretas
                if (
                    hasattr(cell, "column")
                    and cell.column in [5, 9]
                    and self.is_red_cell(cell)
                ):
                    row_num = cell.row
                    if row_num in linhas_faltas:
                        linhas_faltas[row_num] += f", {cell.value}"
                    else:
                        linhas_faltas[row_num] = str(cell.value)

        return linhas_faltas

    def get_sport_default_score(self, sheet_name: str) -> Tuple[int, int]:
        """Retorna o resultado padrão baseado no desporto."""
        sheet_upper = sheet_name.upper()

        if "VOLEIBOL" in sheet_upper:
            return (2, 0)  # 2-0 no voleibol
        elif "FUTSAL" in sheet_upper or "FUTEBOL" in sheet_upper:
            return (3, 0)  # 3-0 nos futebóis
        elif "ANDEBOL" in sheet_upper:
            return (15, 0)  # 15-0 no andebol
        elif "BASQUETEBOL" in sheet_upper:
            return (21, 0)  # 21-0 no basquetebol
        else:
            return (3, 0)  # Padrão para outros desportos

    def apply_default_scores(self, df: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
        """Aplica resultados padrão quando há falta de comparência de apenas uma equipa."""
        if "Falta de Comparência" not in df.columns:
            return df

        # Obtém o resultado padrão para este desporto
        golos_vencedor, golos_perdedor = self.get_sport_default_score(sheet_name)

        for idx, row in df.iterrows():
            falta = row["Falta de Comparência"]

            # Verifica se há falta de comparência e se não está vazio
            if pd.notna(falta) and falta != "":
                # Conta quantas equipas faltaram (separadas por vírgula)
                equipas_faltaram = [
                    equipa.strip() for equipa in falta.split(",") if equipa.strip()
                ]

                # Só aplica resultado se apenas uma equipa faltou
                if len(equipas_faltaram) == 1:
                    equipa_faltou = equipas_faltaram[0]
                    equipa1 = row["Equipa 1"]
                    equipa2 = row["Equipa 2"]
                    # usar o índice diretamente
                    idx_i = idx

                    # Verifica qual equipa faltou e aplica o resultado
                    if equipa_faltou == equipa1:
                        # Equipa 1 faltou, Equipa 2 ganha
                        mask = df.index == idx_i
                        df.loc[mask, "Golos 1"] = golos_perdedor
                        df.loc[mask, "Golos 2"] = golos_vencedor
                    elif equipa_faltou == equipa2:
                        # Equipa 2 faltou, Equipa 1 ganha
                        mask = df.index == idx_i
                        df.loc[mask, "Golos 1"] = golos_vencedor
                        df.loc[mask, "Golos 2"] = golos_perdedor

        return df

    def extract_division_number(self, text: str) -> Optional[str]:
        """Extrai número da divisão do texto."""
        if not isinstance(text, str):
            return None
        match = self.divisao_pattern.search(text)
        return match.group(1) if match else None

    def extract_group(self, text: str) -> Optional[str]:
        """Extrai grupo do texto."""
        if not isinstance(text, str):
            return None
        match = self.grupo_pattern.search(text)
        return match.group(0) if match else None

    def find_divisions_and_groups(
        self, df: pd.DataFrame
    ) -> Tuple[Dict[str, int], Dict[str, int]]:
        """Encontra divisões e grupos no DataFrame."""
        primeira_coluna = df.columns[0]
        divisoes = {}
        grupos = {}

        # Verifica cabeçalho
        num_divisao = self.extract_division_number(primeira_coluna)
        if num_divisao:
            divisoes[num_divisao] = 0

        grupo = self.extract_group(primeira_coluna)
        if grupo:
            grupos[grupo] = 0

        # Percorre dados
        for idx, valor in enumerate(df[primeira_coluna]):
            if isinstance(valor, str) and "DIVISÃO" in valor:
                num_divisao = self.extract_division_number(valor)
                if num_divisao and num_divisao not in divisoes:
                    divisoes[num_divisao] = idx

                grupo = self.extract_group(valor)
                if grupo and grupo not in grupos:
                    grupos[grupo] = idx
            else:
                grupo = self.extract_group(str(valor))
                if grupo and grupo not in grupos:
                    grupos[grupo] = idx

        return divisoes, grupos

    def fill_sections(
        self, df: pd.DataFrame, sections: Dict[str, int], column_name: str
    ) -> pd.DataFrame:
        """Preenche colunas de seção (divisão ou grupo)."""
        if not sections:
            return df

        df[column_name] = ""
        indices = sorted(sections.items(), key=lambda x: x[1])

        for i, (name, start_idx) in enumerate(indices):
            end_idx = indices[i + 1][1] if i + 1 < len(indices) else len(df)

            if column_name == "Grupo":
                clean_name = name.replace("GRUPO ", "").strip()
            else:
                clean_name = name

            df.loc[start_idx : end_idx - 1, column_name] = clean_name

        return df

    def create_headers(self, has_divisions: bool, has_groups: bool) -> List[str]:
        """Cria cabeçalhos baseado na presença de divisões e grupos."""
        headers = self.base_headers.copy()

        if has_divisions:
            headers.append("Divisão")
        if has_groups:
            headers.append("Grupo")

        return headers

    def clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """Limpa e processa o DataFrame."""
        # Remove linhas vazias
        df = df.dropna(how="all").reset_index(drop=True)

        # Filtra linhas válidas (que começam por número ou estão vazias)
        primeira_coluna = df.columns[0]
        df = df[
            df[primeira_coluna].map(lambda x: pd.isna(x) or isinstance(x, (int, float)))
        ]

        # Remove linhas com equipas zeradas
        df = df[~((df["Equipa 1"] == 0) | (df["Equipa 2"] == 0))]

        # Preenche jornadas
        df["Jornada"] = df["Jornada"].ffill()

        # Converte golos para inteiros
        for col in ["Golos 1", "Golos 2"]:
            if col in df.columns:
                df[col] = df[col].apply(lambda x: int(x) if pd.notna(x) else pd.NA)

        return df

    def adjust_journeys(self, df: pd.DataFrame) -> pd.DataFrame:
        """Ajusta números das jornadas para evitar duplicações."""

        def ajustar_jornada(row):
            jornada = row["Jornada"]
            equipa1 = row["Equipa 1"]
            equipa2 = row["Equipa 2"]

            if (jornada, equipa1) in aparicoes or (jornada, equipa2) in aparicoes:
                return jornada + 1

            aparicoes.add((jornada, equipa1))
            aparicoes.add((jornada, equipa2))
            return jornada

        aparicoes = set()
        df["Jornada"] = df.apply(ajustar_jornada, axis=1)
        df["Jornada"] = (
            df["Jornada"].astype(str).str.replace(".1", " (2ª)", regex=False)
        )

        return df

    def sort_by_datetime(self, df: pd.DataFrame) -> pd.DataFrame:
        """Ordena por data e hora, tratando jogos da madrugada (0h-1h) como do dia seguinte."""

        def parse_data_hora(row):
            dia, hora = row["Dia"], row["Hora"]

            if pd.isna(dia):
                return pd.Timestamp.max
            if pd.isna(hora):
                hora = "00:00"

            try:
                data_hora = pd.to_datetime(f"{dia} {hora}", errors="coerce")
                if pd.isna(data_hora):
                    return pd.Timestamp.max

                # Se a hora está entre 0h00 e 0h59, adiciona 24 horas para ordenar depois
                if (
                    data_hora.time() >= pd.to_datetime("00:00").time()
                    and data_hora.time() < pd.to_datetime("01:00").time()
                ):
                    data_hora += pd.Timedelta(hours=24)

                return data_hora
            except Exception:
                return pd.Timestamp.max

        # Coluna auxiliar para data/hora
        df["DataHoraSort"] = df.apply(parse_data_hora, axis=1)

        # Coluna auxiliar para ordenar por jornada quando não há data
        def parse_jornada_sort(val):
            if pd.isna(val):
                return 10**9
            if isinstance(val, (int, float)):
                try:
                    return int(val)
                except Exception:
                    return 10**9
            if isinstance(val, str):
                s = val.strip()
                m = re.match(r"^(\d+)", s)
                if m:
                    try:
                        return int(m.group(1))
                    except Exception:
                        return 10**9
                # Jornadas tipo 'E...' (eliminatórias) vão para o fim do grupo sem data
                return 10**9
            return 10**9

        df["JornadaSort"] = df["Jornada"].apply(parse_jornada_sort)

        # Colunas auxiliares para Divisão e Grupo (caso existam)
        def parse_divisao_sort(val):
            if pd.isna(val):
                return 10**6
            if isinstance(val, (int, float)):
                try:
                    return int(val)
                except Exception:
                    return 10**6
            if isinstance(val, str):
                m = re.search(r"(\d+)", val)
                if m:
                    try:
                        return int(m.group(1))
                    except Exception:
                        return 10**6
            return 10**6

        def parse_grupo_sort(val):
            if pd.isna(val):
                return 10**6
            # Aceita letras (A->1, B->2, ...) ou números diretamente
            if isinstance(val, (int, float)):
                try:
                    return int(val)
                except Exception:
                    return 10**6
            if isinstance(val, str) and val:
                v = val.strip().upper()
                # Se começar por letra
                if v[0].isalpha():
                    return ord(v[0]) - ord("A") + 1
                # Se contiver número
                m = re.search(r"(\d+)", v)
                if m:
                    try:
                        return int(m.group(1))
                    except Exception:
                        return 10**6
            return 10**6

        df["DivisaoSort"] = (
            df["Divisão"].apply(parse_divisao_sort)
            if "Divisão" in df.columns
            else 10**6
        )
        df["GrupoSort"] = (
            df["Grupo"].apply(parse_grupo_sort) if "Grupo" in df.columns else 10**6
        )

        # Ordenar por data/hora, depois jornada, depois Divisão e Grupo, e por fim nomes de equipas
        df = df.sort_values(
            [
                "DataHoraSort",
                "JornadaSort",
                "DivisaoSort",
                "GrupoSort",
                "Equipa 1",
                "Equipa 2",
            ],
            ascending=[True, True, True, True, True, True],
        )
        df = df.drop(
            columns=["DataHoraSort", "JornadaSort", "DivisaoSort", "GrupoSort"]
        )

        return df

    def finalize_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """Finalizações finais do DataFrame."""
        # Remove linhas vazias nas colunas principais
        colunas_principais = [
            "Dia",
            "Hora",
            "Local",
            "Equipa 1",
            "Golos 1",
            "Golos 2",
            "Equipa 2",
        ]
        df = df.dropna(subset=colunas_principais, how="all")

        # Garantir coluna "Falta de Comparência" presente e no fim, mesmo vazia
        if "Falta de Comparência" not in df.columns:
            df["Falta de Comparência"] = ""

        colunas = [col for col in df.columns if col != "Falta de Comparência"]
        colunas.append("Falta de Comparência")
        df = df[colunas]

        return df

    def process_sheet(self, sheet_name: str) -> bool:
        """Processa uma folha específica."""
        if sheet_name == "CASTIGOS":
            return False

        print(f"A processar a folha: {sheet_name}")

        # Verifica se tem divisões
        sample = self.xls.parse(sheet_name, nrows=1)
        has_div = "DIVISÃO" in str(sample.columns[0])

        # Extrai células vermelhas
        linhas_faltas = self.extract_red_cells(sheet_name)

        # Carrega dados
        df = pd.read_excel(
            self.xls, sheet_name=sheet_name, usecols=[0, 1, 2, 3, 4, 5, 7, 8]
        )
        df["Falta de Comparência"] = ""

        # Preenche faltas de comparência
        for row_idx, value in linhas_faltas.items():
            df_idx = row_idx - 2  # Ajusta índice
            if 0 <= df_idx < len(df):
                df.at[df_idx, "Falta de Comparência"] = value

        # Limpa DataFrame inicial
        df = df.dropna(how="all").reset_index(drop=True)

        # Processa divisões e grupos
        if has_div:
            divisoes, grupos = self.find_divisions_and_groups(df)
            df = self.fill_sections(df, divisoes, "Divisão")
            df = self.fill_sections(df, grupos, "Grupo")
            headers = self.create_headers(bool(divisoes), bool(grupos))
        else:
            _, grupos = self.find_divisions_and_groups(df)
            df = self.fill_sections(df, grupos, "Grupo")
            headers = self.create_headers(False, bool(grupos))

        # Ajusta cabeçalhos
        df = df.iloc[:, : len(headers)]
        df.columns = headers

        # Processa dados
        df = self.clean_dataframe(df)
        df = self.adjust_journeys(df)
        df = self.sort_by_datetime(df)
        df = self.apply_default_scores(df, sheet_name)
        df = self.finalize_dataframe(df)

        # Não adicionar coluna de Época nos CSVs

        # Salva resultado
        if self.season:
            filename = f"{sheet_name}_{self.season}.csv"
        else:
            filename = f"{sheet_name}.csv"

        output_file = self.output_dir / filename
        df.to_csv(output_file, index=False)

        print(f"Folha '{sheet_name}' processada e salva em '{output_file}'")
        return True

    def process_all_sheets(self):
        """Processa todas as folhas do Excel."""
        processed_count = 0
        target_sheets = (
            self._sheets_to_process if self._sheets_to_process else self.xls.sheet_names
        )
        for sheet_name in target_sheets:
            if self.process_sheet(str(sheet_name)):
                processed_count += 1

        print(f"\nProcessamento concluído! {processed_count} folhas processadas.")


def main():
    """Função principal."""
    # 1) Ler config/env para obter a URL do documento de resultados
    config_url: Optional[str] = None
    config_path = Path("config.json")
    if config_path.exists():
        try:
            with open(config_path, "r", encoding="utf-8") as f:
                cfg = json.load(f)
                config_url = cfg.get("results_url")
        except Exception as e:
            print(f"Aviso: Não foi possível ler config.json: {e}")

    # Variável de ambiente como fallback
    env_url = os.environ.get("RESULTS_URL")
    if not config_url and env_url:
        config_url = env_url

    downloaded_file: Optional[Path] = None
    season_detected: Optional[str] = None

    # 2) Se houver URL, descarregar o ficheiro atualizado
    if config_url:
        try:
            url = normalize_results_url(config_url)
            downloaded_file = download_results_excel(url)
            print(f"Documento descarregado para: {downloaded_file}")

            # Abrir para detetar época mais recente a partir das folhas
            xls_temp = pd.ExcelFile(str(downloaded_file))
            season_detected = detect_latest_season_from_sheet_names(
                list(map(str, xls_temp.sheet_names))
            )
            # Fechar o handler do Excel antes de renomear em Windows
            try:
                xls_temp.close()
            except Exception:
                pass

            # Se detetou época, renomear o ficheiro local com a época
            if season_detected:
                target_name = f"Resultados Taça UA {season_detected}.xlsx"
            else:
                # fallback se não houver época nas folhas -> tentar a partir do nome
                extracted = extract_season_from_filename(downloaded_file.name)
                if not extracted:
                    extracted = current_season_token()
                target_name = f"Resultados Taça UA {extracted}.xlsx"

            target_path = downloaded_file.parent / target_name
            if downloaded_file.name != target_name:
                try:
                    # Substitui se já existir
                    if target_path.exists():
                        target_path.unlink()
                    downloaded_file.rename(target_path)
                    downloaded_file = target_path
                except Exception as e:
                    print(
                        f"Aviso: Não foi possível renomear o ficheiro descarregado: {e}"
                    )

        except Exception as e:
            print(f"Erro ao descarregar o documento: {e}")
            downloaded_file = None

    # 3) Determinar o caminho do ficheiro a processar
    if downloaded_file and downloaded_file.exists():
        file_path = str(downloaded_file)
    else:
        # fallback: usar ficheiro local existente (compatibilidade antiga)
        # tentar usar padrão com época corrente
        default_local = f"Resultados Taça UA {current_season_token()}.xlsx"
        if not os.path.exists(default_local):
            print("Erro: Nenhum ficheiro local encontrado e URL não disponível/valida.")
            print(
                "Defina 'results_url' em config.json ou a variável de ambiente RESULTS_URL."
            )
            return
        file_path = default_local

    # 4) Preparar backup e verificação de mudanças
    # Tenta inferir época para nome do backup
    season_for_backup = (
        season_detected
        or extract_season_from_filename(Path(file_path).name)
        or "latest"
    )
    backup_file = f"backup_Resultados Taça UA {season_for_backup}.xlsx"

    if os.path.exists(backup_file) and files_are_identical(file_path, backup_file):
        print(
            "O arquivo Excel não mudou desde a última execução. Nenhum processamento necessário."
        )
        return

    print("Arquivo Excel mudou ou primeira execução. Processando dados...")

    try:
        shutil.copy2(file_path, backup_file)
        print(f"Backup criado: {backup_file}")
    except Exception as e:
        print(f"Aviso: Não foi possível criar backup: {e}")

    # 5) Selecionar folhas a processar com base na época
    xls_all = pd.ExcelFile(file_path)
    if not season_detected:
        season_detected = detect_latest_season_from_sheet_names(
            list(map(str, xls_all.sheet_names))
        )

    sheets_to_process = choose_sheets_for_season(
        list(map(str, xls_all.sheet_names)), season_detected
    )

    # 6) Processar o ficheiro
    processor = ExcelProcessor(
        file_path, season_override=season_detected, sheets_to_process=sheets_to_process
    )
    processor.process_all_sheets()


if __name__ == "__main__":
    main()
